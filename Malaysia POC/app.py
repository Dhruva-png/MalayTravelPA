"""
MarvelAI Travel PA Insurance Claims Portal
==========================================

requirements.txt:
-----------------
streamlit>=1.32.0
plotly>=5.20.0
pandas>=2.2.0
requests>=2.31.0
Pillow>=10.2.0
PyMuPDF>=1.23.0
# Optional – only needed for image/scanned-PDF OCR:
# pytesseract>=0.3.10

Usage:
------
1. Install dependencies:  pip install -r requirements.txt
2. Ensure Ollama is running locally with llama3.2 pulled:
       ollama pull llama3.2
       ollama serve
3. Run the app:  streamlit run app.py
"""

# ──────────────────────────────────────────────────────────────────────────────
# IMPORTS
# ──────────────────────────────────────────────────────────────────────────────
import json
import hashlib
import random
import re
import time
import uuid
from datetime import date, datetime, timedelta
from io import BytesIO
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import requests
import streamlit as st
import streamlit.components.v1

# ──────────────────────────────────────────────────────────────────────────────
# PAGE CONFIG  (must be first Streamlit call)
# ──────────────────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="MarvelAI · Claims Portal",
    page_icon="🛡️",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ──────────────────────────────────────────────────────────────────────────────
# GLOBAL CONSTANTS & CONFIG
# ──────────────────────────────────────────────────────────────────────────────
OLLAMA_HOST  = "http://localhost:11434"
OLLAMA_URL   = f"{OLLAMA_HOST}/api/generate"
OLLAMA_TAGS_URL = f"{OLLAMA_HOST}/api/tags"
OLLAMA_MODEL = "llama3.2"
OLLAMA_FALLBACK_MODELS = ["llama3.1"]
OLLAMA_TIMEOUT = 120          # seconds

COVERAGE_TYPES = [
    "Accidental Death", "Permanent Disability", "Hospitalisation",
    "Emergency Evacuation", "Trip Cancellation", "Baggage Loss",
]

GEOGRAPHIES = [
    "Southeast Asia", "Europe", "North America",
    "Middle East", "South Asia", "Oceania", "Africa",
]

DOC_TYPES = [
    "Death Certificate", "Hospital Bill", "Medical Report",
    "Disability Assessment", "Police Report", "Boarding Pass",
    "Insurance Policy", "Claim Form", "Discharge Summary", "Unknown",
]

# ──────────────────────────────────────────────────────────────────────────────
# CUSTOM CSS  – professional light theme
# ──────────────────────────────────────────────────────────────────────────────
def inject_css() -> None:
    st.markdown("""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap');

    /* ── Design tokens ──────────────────────────────────────────────────────── */
    :root {
        /* Brand */
        --navy          : #0D1F3C;
        --navy-mid      : #1A3354;
        --blue          : #1A6FE0;
        --blue-light    : #E8F1FC;

        /* Semantic */
        --success       : #0A7C5C;
        --success-bg    : #E6F5F0;
        --warning       : #7A4F00;
        --warning-bg    : #FEF3CD;
        --danger        : #B91C1C;
        --danger-bg     : #FEE8E8;
        --info          : #1753A8;
        --info-bg       : #EBF2FD;

        /* Surfaces */
        --surface       : #FFFFFF;
        --surface-page  : #F0F4F9;
        --surface-raised: #FFFFFF;
        --border        : #D8E1EE;
        --border-light  : #EAF0F8;

        /* Text — all on white or #F0F4F9 backgrounds */
        --text-heading  : #0D1F3C;
        --text-body     : #243550;
        --text-muted    : #5A6E89;
        --text-faint    : #8FA3BC;

        /* Typography */
        --font          : "Inter", "Segoe UI", system-ui, sans-serif;
        --font-mono     : "JetBrains Mono", "Fira Code", "Consolas", monospace;
    }

    /* ── Global base ─────────────────────────────────────────────────────────  */
    html, body,
    [data-testid="stAppViewContainer"],
    [data-testid="stMain"],
    .main .block-container {
        background-color: var(--surface-page) !important;
        font-family: var(--font) !important;
        color: var(--text-body) !important;
    }

    /* Streamlit default text overrides */
    p, li, span, label, div {
        color: var(--text-body);
    }
    h1, h2, h3, h4 {
        color: var(--text-heading) !important;
        font-family: var(--font) !important;
    }

    /* ── Sidebar ─────────────────────────────────────────────────────────────  */
    [data-testid="stSidebar"] {
        background-color: var(--navy) !important;
        border-right: 2px solid var(--navy-mid) !important;
        padding-top: 0 !important;
    }
    /* All text inside sidebar: white */
    [data-testid="stSidebar"],
    [data-testid="stSidebar"] p,
    [data-testid="stSidebar"] span,
    [data-testid="stSidebar"] div,
    [data-testid="stSidebar"] label {
        color: #FFFFFF !important;
    }
    /* Radio option labels specifically */
    [data-testid="stSidebar"] .stRadio > div > label {
        color: rgba(255,255,255,0.82) !important;
        font-size: 0.875rem;
        padding: 0.45rem 0.6rem;
        border-radius: 6px;
        margin: 1px 0;
        transition: background 0.15s, color 0.15s;
        display: block;
    }
    [data-testid="stSidebar"] .stRadio > div > label:hover {
        background: rgba(255,255,255,0.08) !important;
        color: #FFFFFF !important;
    }
    /* Selected radio item */
    [data-testid="stSidebar"] .stRadio > div > label[data-baseweb="radio"] > div:first-child {
        border-color: var(--blue) !important;
        background-color: var(--blue) !important;
    }
    /* Sidebar buttons */
    [data-testid="stSidebar"] .stButton > button {
        background: rgba(255,255,255,0.08) !important;
        color: rgba(255,255,255,0.75) !important;
        border: 1px solid rgba(255,255,255,0.15) !important;
        border-radius: 6px;
        font-size: 0.82rem;
        transition: background 0.15s;
    }
    [data-testid="stSidebar"] .stButton > button:hover {
        background: rgba(255,255,255,0.14) !important;
        color: #FFFFFF !important;
    }
    /* Sidebar divider */
    [data-testid="stSidebar"] hr {
        border-color: rgba(255,255,255,0.12) !important;
    }

    /* ── Main content padding ────────────────────────────────────────────────  */
    .main .block-container {
        padding: 2rem 2.5rem 3rem !important;
        max-width: 1400px;
    }

    /* ── Page title & subtitle ───────────────────────────────────────────────  */
    [data-testid="stMarkdownContainer"] h2 {
        font-size: 1.45rem;
        font-weight: 700;
        color: var(--text-heading) !important;
        margin-bottom: 0.25rem;
        letter-spacing: -0.02em;
    }
    [data-testid="stMarkdownContainer"] > p:first-of-type {
        color: var(--text-muted) !important;
        font-size: 0.9rem;
        margin-top: 0;
    }

    /* ── Cards ───────────────────────────────────────────────────────────────  */
    .card {
        background: var(--surface);
        border: 1px solid var(--border);
        border-radius: 10px;
        padding: 1.35rem 1.5rem;
        margin-bottom: 1rem;
    }
    .card h3 {
        font-size: 1rem;
        font-weight: 600;
        color: var(--text-heading);
        margin: 0 0 0.25rem;
    }

    /* ── Metric cards ────────────────────────────────────────────────────────  */
    .card-metric {
        background: var(--surface);
        border: 1px solid var(--border);
        border-radius: 10px;
        padding: 1.1rem 1rem 1rem;
        text-align: center;
    }
    .metric-value {
        font-size: 1.9rem;
        font-weight: 700;
        color: var(--text-heading);
        line-height: 1.1;
        letter-spacing: -0.02em;
    }
    .metric-label {
        font-size: 0.7rem;
        color: var(--text-muted);
        text-transform: uppercase;
        letter-spacing: 0.09em;
        margin-top: 0.35rem;
        font-weight: 500;
    }
    .metric-delta-up   { color: var(--success); font-size: 0.78rem; margin-top: 0.2rem; }
    .metric-delta-down { color: var(--danger);  font-size: 0.78rem; margin-top: 0.2rem; }

    /* ── Section headers ─────────────────────────────────────────────────────  */
    .section-header {
        font-size: 0.78rem;
        font-weight: 700;
        color: var(--text-muted);
        text-transform: uppercase;
        letter-spacing: 0.1em;
        border-bottom: 2px solid var(--border-light);
        padding-bottom: 0.45rem;
        margin: 1.5rem 0 0.9rem;
    }

    /* ── Status badges ───────────────────────────────────────────────────────  */
    .badge {
        display: inline-block;
        padding: 0.18rem 0.6rem;
        border-radius: 4px;
        font-size: 0.7rem;
        font-weight: 600;
        letter-spacing: 0.04em;
        text-transform: uppercase;
        line-height: 1.5;
    }
    .badge-success { background: var(--success-bg); color: var(--success); }
    .badge-warning { background: var(--warning-bg); color: var(--warning); }
    .badge-danger  { background: var(--danger-bg);  color: var(--danger);  }
    .badge-info    { background: var(--info-bg);    color: var(--info);    }
    .badge-neutral { background: var(--border-light); color: var(--text-muted); }

    /* ── Fraud score bar ─────────────────────────────────────────────────────  */
    .fraud-bar-wrapper {
        width: 100%;
        background: var(--border-light);
        border-radius: 999px;
        height: 8px;
        margin: 0.5rem 0 0.25rem;
        overflow: hidden;
    }
    .fraud-bar {
        height: 8px;
        border-radius: 999px;
    }
    .fraud-score-label {
        font-size: 0.75rem;
        color: var(--text-muted);
        margin-top: 0.15rem;
    }

    /* ── JSON box ────────────────────────────────────────────────────────────  */
    .json-box {
        background: var(--surface-page);
        border: 1px solid var(--border);
        border-radius: 8px;
        padding: 0.9rem 1rem;
        font-family: var(--font-mono);
        font-size: 0.78rem;
        color: var(--text-body);
        overflow-x: auto;
        line-height: 1.75;
    }

    /* ── Streamlit widget label text ─────────────────────────────────────────  */
    .stSelectbox label,
    .stMultiSelect label,
    .stFileUploader label,
    .stDateInput label,
    .stTextInput label,
    .stExpander label,
    [data-testid="stWidgetLabel"] {
        color: var(--text-body) !important;
        font-size: 0.85rem !important;
        font-weight: 500 !important;
    }

    /* ── Expander ────────────────────────────────────────────────────────────  */
    [data-testid="stExpander"] {
        border: 1px solid var(--border) !important;
        border-radius: 8px !important;
        background: var(--surface) !important;
    }
    [data-testid="stExpander"] summary {
        color: var(--text-body) !important;
        font-weight: 500;
        font-size: 0.88rem;
    }

    /* ── Info / warning / error / success banners ────────────────────────────  */
    [data-testid="stAlert"] {
        border-radius: 8px !important;
        font-size: 0.875rem;
    }

    /* ── Dataframe / table header ────────────────────────────────────────────  */
    [data-testid="stDataFrame"] thead th,
    .dataframe thead th {
        background-color: var(--navy) !important;
        color: #FFFFFF !important;
        font-size: 0.72rem !important;
        font-weight: 600 !important;
        text-transform: uppercase;
        letter-spacing: 0.07em;
        padding: 0.5rem 0.75rem !important;
    }
    [data-testid="stDataFrame"] tbody tr:nth-child(even) {
        background-color: var(--surface-page) !important;
    }
    [data-testid="stDataFrame"] tbody td {
        color: var(--text-body) !important;
        font-size: 0.83rem !important;
    }

    /* ── Buttons (main area) ─────────────────────────────────────────────────  */
    .stButton > button {
        background: var(--blue) !important;
        color: #FFFFFF !important;
        border: none !important;
        border-radius: 6px !important;
        font-size: 0.85rem !important;
        font-weight: 500 !important;
        padding: 0.45rem 1.1rem !important;
        transition: opacity 0.15s;
    }
    .stButton > button:hover {
        opacity: 0.88 !important;
    }

    /* ── Multiselect tags ────────────────────────────────────────────────────  */
    [data-testid="stMultiSelect"] span[data-baseweb="tag"] {
        background: var(--blue-light) !important;
        color: var(--blue) !important;
        border-radius: 4px !important;
        font-size: 0.78rem !important;
    }

    /* ── Spinner text ────────────────────────────────────────────────────────  */
    [data-testid="stSpinner"] p {
        color: var(--text-muted) !important;
        font-size: 0.85rem;
    }

    /* ── Caption text ────────────────────────────────────────────────────────  */
    [data-testid="stCaptionContainer"] p,
    .stCaption {
        color: var(--text-muted) !important;
        font-size: 0.8rem !important;
    }

    /* ── Code blocks ─────────────────────────────────────────────────────────  */
    code, pre {
        font-family: var(--font-mono) !important;
        font-size: 0.8rem;
        background: var(--surface-page);
        color: var(--text-body);
        border-radius: 4px;
    }

    /* ── Divider ─────────────────────────────────────────────────────────────  */
    hr {
        border-color: var(--border-light) !important;
        margin: 0.75rem 0 !important;
    }

    /* ── Hide Streamlit chrome ───────────────────────────────────────────────  */
    #MainMenu, footer, header { visibility: hidden; }
    [data-testid="stDecoration"] { display: none; }

    /* Professional refresh layer */
    :root {
        --navy: #101828;
        --navy-mid: #24324A;
        --blue: #2563EB;
        --blue-light: #EFF6FF;
        --surface-page: #F6F8FB;
        --surface: #FFFFFF;
        --border: #E3E8F0;
        --border-light: #EEF2F7;
        --text-heading: #111827;
        --text-body: #344054;
        --text-muted: #667085;
        --success: #067647;
        --success-bg: #ECFDF3;
        --warning: #B54708;
        --warning-bg: #FFFAEB;
        --danger: #B42318;
        --danger-bg: #FEF3F2;
    }
    .main .block-container {
        padding: 1.5rem 2rem 3rem !important;
    }
    [data-testid="stSidebar"] {
        background: linear-gradient(180deg, #111827 0%, #172033 100%) !important;
        border-right: 1px solid rgba(255,255,255,0.08) !important;
    }
    [data-testid="stSidebar"] .stRadio > div {
        gap: 0.25rem;
    }
    [data-testid="stSidebar"] .stRadio > div > label {
        border-radius: 8px !important;
        padding: 0.62rem 0.7rem !important;
    }
    [data-testid="stSidebar"] .stRadio > div > label:hover {
        background: rgba(255,255,255,0.10) !important;
    }
    .page-hero {
        background: linear-gradient(135deg, #FFFFFF 0%, #F4F7FB 56%, #EAF2FF 100%);
        border: 1px solid var(--border);
        border-radius: 8px;
        padding: 1.35rem 1.5rem;
        margin: 0 0 1.25rem;
        box-shadow: 0 14px 38px rgba(16, 24, 40, 0.06);
    }
    .page-kicker {
        color: var(--blue);
        font-size: 0.72rem;
        font-weight: 700;
        letter-spacing: 0.11em;
        text-transform: uppercase;
        margin-bottom: 0.4rem;
    }
    .page-hero h1 {
        color: var(--text-heading) !important;
        font-size: 1.85rem;
        line-height: 1.15;
        margin: 0 0 0.35rem;
        letter-spacing: 0;
    }
    .page-hero p {
        color: var(--text-muted) !important;
        font-size: 0.95rem;
        margin: 0;
        max-width: 760px;
    }
    .card, .card-metric, [data-testid="stExpander"] {
        border-radius: 8px !important;
        border: 1px solid var(--border) !important;
        box-shadow: 0 8px 24px rgba(16, 24, 40, 0.045);
    }
    .card-metric {
        min-height: 116px;
        display: flex;
        flex-direction: column;
        justify-content: center;
    }
    .metric-value {
        font-size: 2rem;
        letter-spacing: 0;
    }
    .metric-label, .section-header {
        letter-spacing: 0.08em;
    }
    .section-header {
        border-bottom: 1px solid var(--border);
        color: var(--text-heading);
        margin-top: 1.35rem;
    }
    [data-testid="stPlotlyChart"] {
        background: #FFFFFF;
        border: 1px solid var(--border);
        border-radius: 8px;
        padding: 0.5rem;
        box-shadow: 0 8px 24px rgba(16, 24, 40, 0.045);
    }
    .stButton > button {
        border-radius: 8px !important;
        min-height: 2.45rem;
        box-shadow: 0 1px 2px rgba(16, 24, 40, 0.08);
    }
    [data-testid="stFileUploader"] section {
        border: 1px dashed #98A2B3 !important;
        border-radius: 8px !important;
        background: #FFFFFF !important;
    }
    /* Final art direction: executive claims console */
    :root {
        --ink: #172033;
        --ink-soft: #344054;
        --canvas: #F7F5F0;
        --panel: #FFFFFF;
        --panel-soft: #FBFAF7;
        --line: #E4DFD6;
        --line-strong: #D6CEC0;
        --brand: #1F6F68;
        --brand-deep: #174E4A;
        --brand-soft: #E8F3F1;
        --accent: #C98A2E;
        --accent-soft: #FFF4E4;
        --blue: #2E5EAA;
        --blue-light: #E9F0FA;
        --success: #1F7A4D;
        --success-bg: #EAF6EF;
        --warning: #A86512;
        --warning-bg: #FFF4E4;
        --danger: #B2382F;
        --danger-bg: #FCEDEA;
        --text-heading: var(--ink);
        --text-body: var(--ink-soft);
        --text-muted: #716B62;
        --surface-page: var(--canvas);
        --surface: var(--panel);
        --border: var(--line);
        --border-light: #EEE9E0;
    }
    html, body,
    [data-testid="stAppViewContainer"],
    [data-testid="stMain"],
    .main .block-container {
        background:
            radial-gradient(circle at top left, rgba(31,111,104,0.08), transparent 32rem),
            linear-gradient(180deg, #FBFAF7 0%, var(--canvas) 52%, #F1EEE7 100%) !important;
    }
    .main .block-container {
        max-width: 1320px;
        padding: 1.65rem 2.1rem 3.25rem !important;
    }
    [data-testid="stSidebar"] {
        background:
            linear-gradient(180deg, #172033 0%, #142A2E 58%, #102020 100%) !important;
        border-right: 1px solid rgba(255,255,255,0.10) !important;
        box-shadow: 16px 0 38px rgba(23,32,51,0.14);
    }
    [data-testid="stSidebar"] [data-testid="stVerticalBlock"] {
        gap: 0.35rem;
    }
    [data-testid="stSidebar"] .stRadio > div {
        gap: 0.35rem;
    }
    [data-testid="stSidebar"] .stRadio > div > label {
        background: transparent !important;
        border: 1px solid transparent;
        border-radius: 8px !important;
        color: rgba(255,255,255,0.78) !important;
        padding: 0.72rem 0.78rem !important;
        font-weight: 600;
        display: flex !important;
        align-items: center !important;
        gap: 0.65rem !important;
        min-height: 2.55rem;
    }
    [data-testid="stSidebar"] .stRadio > div > label > div:first-child {
        display: flex !important;
        align-items: center !important;
        justify-content: center !important;
        margin-top: 0 !important;
    }
    [data-testid="stSidebar"] .stRadio input {
        margin-top: 0 !important;
    }
    [data-testid="stSidebar"] .stRadio > div > label:hover {
        background: rgba(255,255,255,0.08) !important;
        border-color: rgba(255,255,255,0.08);
        color: #FFFFFF !important;
    }
    [data-testid="stSidebar"] .stRadio > div > label:has(input:checked) {
        background: rgba(255,255,255,0.13) !important;
        border-color: rgba(255,255,255,0.18);
        box-shadow: inset 3px 0 0 var(--accent);
    }
    .page-hero {
        position: relative;
        overflow: hidden;
        background:
            linear-gradient(135deg, rgba(255,255,255,0.98) 0%, rgba(251,250,247,0.98) 56%, rgba(232,243,241,0.98) 100%);
        border: 1px solid rgba(214,206,192,0.95);
        border-radius: 8px;
        padding: 1.55rem 1.75rem;
        margin: 0 0 1.35rem;
        box-shadow: 0 18px 45px rgba(23,32,51,0.08);
    }
    .page-hero::after {
        content: "";
        position: absolute;
        inset: auto 1.4rem 0 1.4rem;
        height: 3px;
        background: linear-gradient(90deg, var(--brand), var(--accent), transparent);
        opacity: 0.9;
    }
    .page-kicker {
        color: var(--brand-deep);
        font-size: 0.7rem;
        font-weight: 800;
        letter-spacing: 0.12em;
        text-transform: uppercase;
        margin-bottom: 0.5rem;
    }
    .page-hero h1 {
        color: var(--ink) !important;
        font-size: 2rem;
        font-weight: 750;
        line-height: 1.12;
        margin: 0 0 0.4rem;
        letter-spacing: 0;
    }
    .page-hero p {
        color: #5F6673 !important;
        font-size: 0.98rem;
        line-height: 1.55;
        max-width: 820px;
    }
    .card,
    .card-metric,
    [data-testid="stExpander"],
    [data-testid="stPlotlyChart"],
    [data-testid="stDataFrame"] {
        background: rgba(255,255,255,0.94) !important;
        border: 1px solid rgba(214,206,192,0.88) !important;
        border-radius: 8px !important;
        box-shadow: 0 12px 30px rgba(23,32,51,0.055);
    }
    .card {
        padding: 1.25rem 1.35rem;
    }
    .card-metric {
        min-height: 118px;
        align-items: flex-start;
        text-align: left;
        padding: 1.15rem 1.2rem;
        background:
            linear-gradient(180deg, #FFFFFF 0%, #FBFAF7 100%) !important;
    }
    .metric-value {
        color: var(--ink);
        font-size: 2.05rem;
        font-weight: 760;
    }
    .metric-label {
        color: #7A746B;
        font-size: 0.68rem;
        font-weight: 800;
        letter-spacing: 0.09em;
    }
    .metric-delta-up,
    .metric-delta-down {
        border-radius: 999px;
        padding: 0.12rem 0.5rem;
        font-weight: 700;
    }
    .metric-delta-up {
        background: var(--success-bg);
        color: var(--success);
    }
    .metric-delta-down {
        background: var(--danger-bg);
        color: var(--danger);
    }
    .section-header {
        border-bottom: 1px solid var(--line);
        color: var(--ink);
        font-size: 0.72rem;
        font-weight: 850;
        letter-spacing: 0.11em;
        margin: 1.45rem 0 0.95rem;
        padding-bottom: 0.55rem;
    }
    .section-header::before {
        content: "";
        display: inline-block;
        width: 8px;
        height: 8px;
        margin-right: 0.45rem;
        border-radius: 50%;
        background: var(--accent);
    }
    .badge {
        border-radius: 999px;
        padding: 0.22rem 0.62rem;
        font-size: 0.68rem;
        font-weight: 800;
    }
    .badge-info,
    .badge-neutral {
        background: var(--brand-soft);
        color: var(--brand-deep);
    }
    .stButton > button {
        background: linear-gradient(180deg, var(--brand) 0%, var(--brand-deep) 100%) !important;
        border: 1px solid rgba(16,80,74,0.15) !important;
        border-radius: 8px !important;
        color: #FFFFFF !important;
        font-weight: 750 !important;
        min-height: 2.55rem;
        box-shadow: 0 8px 18px rgba(31,111,104,0.18);
    }
    .stButton > button:hover {
        box-shadow: 0 10px 22px rgba(31,111,104,0.24);
        transform: translateY(-1px);
        opacity: 1 !important;
    }
    [data-testid="stFileUploader"] section,
    [data-baseweb="select"] > div,
    [data-baseweb="input"] {
        background: #FFFFFF !important;
        border-color: var(--line-strong) !important;
        border-radius: 8px !important;
    }
    [data-testid="stExpander"] summary {
        color: var(--ink) !important;
        font-weight: 750;
    }
    [data-testid="stPlotlyChart"] {
        padding: 0.75rem;
        overflow: hidden;
    }
    [data-testid="stPlotlyChart"] > div {
        overflow: hidden;
    }
    [data-testid="stDataFrame"] thead th,
    .dataframe thead th {
        background: #F0EAE0 !important;
        color: var(--ink) !important;
    }
    [data-testid="stDataFrame"],
    [data-testid="stDataFrame"] div,
    [data-testid="stDataFrame"] canvas,
    [data-testid="stDataFrame"] iframe {
        background-color: #FFFFFF !important;
        color: var(--ink-soft) !important;
    }
    [data-testid="stDataFrame"] tbody tr,
    [data-testid="stDataFrame"] tbody td {
        background-color: #FFFFFF !important;
        color: var(--ink-soft) !important;
    }
    [data-testid="stDataFrame"] tbody tr:nth-child(even),
    [data-testid="stDataFrame"] [role="row"]:nth-child(even) {
        background-color: #FBFAF7 !important;
    }
    [data-testid="stAlert"] {
        border-radius: 8px !important;
        border: 1px solid var(--line) !important;
        box-shadow: 0 8px 22px rgba(23,32,51,0.04);
    }
    .json-box,
    code,
    pre {
        background: #F4F0E8 !important;
        border-color: var(--line) !important;
    }
    @media (max-width: 900px) {
        .main .block-container { padding: 1rem 1rem 2rem !important; }
        .page-hero h1 { font-size: 1.45rem; }
        .page-hero { padding: 1rem; }
    }
    </style>
    """, unsafe_allow_html=True)


# ──────────────────────────────────────────────────────────────────────────────
# MOCK DATA GENERATORS
# ──────────────────────────────────────────────────────────────────────────────
def generate_mock_claims(n: int = 120) -> pd.DataFrame:
    """Generate synthetic claims data for dashboard analytics."""
    random.seed(42)
    statuses  = ["Approved", "Pending", "Rejected", "Under Review", "Fraud Flagged"]
    weights   = [0.40, 0.25, 0.15, 0.12, 0.08]
    base_date = datetime(2024, 1, 1)

    records = []
    for i in range(n):
        cov  = random.choice(COVERAGE_TYPES)
        geo  = random.choice(GEOGRAPHIES)
        stat = random.choices(statuses, weights=weights)[0]
        fraud_score = (
            random.uniform(65, 98) if stat == "Fraud Flagged"
            else random.uniform(0, 40)
        )
        amount_map = {
            "Accidental Death"    : (50000, 500000),
            "Permanent Disability": (20000, 300000),
            "Hospitalisation"     : (2000,  80000),
            "Emergency Evacuation": (5000,  150000),
            "Trip Cancellation"   : (500,   15000),
            "Baggage Loss"        : (200,   5000),
        }
        lo, hi = amount_map[cov]
        records.append({
            "claim_id"     : f"CLM-{10000 + i}",
            "submitted_at" : base_date + timedelta(days=random.randint(0, 365)),
            "coverage_type": cov,
            "geography"    : geo,
            "status"       : stat,
            "claim_amount" : round(random.uniform(lo, hi), 2),
            "fraud_score"  : round(fraud_score, 1),
            "processing_days": random.randint(1, 45),
        })
    return pd.DataFrame(records)


def mock_ocr_text(doc_type_hint: str = "hospital_bill") -> str:
    """Return mock OCR-extracted text for a given document type."""
    samples = {
        "hospital_bill": """
CITY GENERAL HOSPITAL — PATIENT BILL
Patient: Rajesh Kumar Sharma
Policy No: TRV-2024-INS-88742
Date of Admission: 14 March 2024    Date of Discharge: 19 March 2024
Accident Date: 13 March 2024        Location: Bangkok, Thailand
Diagnosis: Fracture of right radius (S52.3), Laceration of forearm
Treatment: Surgical fixation, physiotherapy sessions
Room Charges: ₹42,000   Surgical Charges: ₹1,18,500
Medicines: ₹12,340      Consultation: ₹8,500
TOTAL DUE: ₹1,81,340
Nominee: Priya Sharma (Spouse)
Bank Account: HDFC Bank, A/C 50200012345678, IFSC HDFC0001234
""",
        "death_certificate": """
DEATH CERTIFICATE — MUNICIPALITY OF PARIS, FRANCE
Deceased: Anita Raghunathan, DOB: 12-Jun-1978
Date of Death: 03 November 2024
Cause: Road Traffic Accident — multiple trauma
Policy No: TRV-2024-GBL-44211
Insurer: MarvelAI Travel PA
Nominee/Claimant: Vikram Raghunathan (Husband)
Nationality: Indian    Passport: P9234567
Bank Account: SBI, A/C 31209876543, IFSC SBIN0007654
""",
        "disability_report": """
DISABILITY ASSESSMENT REPORT
Patient: Santhosh Pillai     Policy No: TRV-2024-SEA-30019
Date of Accident: 22 January 2024    Location: Kuala Lumpur, Malaysia
Nature of Injury: Traumatic amputation of left index finger (M00.87)
Disability %: 15% Permanent Partial Disability
Assessed by: Dr. Fatimah Binti Hassan, MBBS, MS Ortho
Hospital: Sunway Medical Centre, Petaling Jaya
Hospitalisation: 22 Jan 2024 – 28 Jan 2024
Treatment Cost: MYR 28,500 (approx ₹5,10,000)
Nominee: Deepa Pillai (Spouse)
Bank Account: Axis Bank, A/C 9170200012345678, IFSC UTIB0001234
""",
    }
    return samples.get(doc_type_hint, samples["hospital_bill"])


# ──────────────────────────────────────────────────────────────────────────────
# OLLAMA LLM INTEGRATION
# ──────────────────────────────────────────────────────────────────────────────
def _ocr_image_bytes(image_bytes: bytes) -> Tuple[str, str]:
    """Extract text from an image when local OCR support is installed."""
    try:
        from PIL import Image
        import pytesseract
    except ImportError:
        return "", "Image OCR requires `pytesseract` and `Pillow`."

    try:
        image = Image.open(BytesIO(image_bytes))
        text = pytesseract.image_to_string(image).strip()
        return text, ""
    except Exception as exc:
        return "", f"Image OCR failed: {exc}"


def extract_pdf_text(file_bytes: bytes) -> Tuple[str, str, str]:
    """Extract selectable PDF text and OCR rendered pages when needed."""
    try:
        import fitz
    except ImportError:
        return "", "PDF text extraction requires `PyMuPDF`.", "PDF extraction unavailable"

    try:
        doc = fitz.open(stream=file_bytes, filetype="pdf")
    except Exception as exc:
        return "", f"Could not open PDF: {exc}", "PDF open failed"

    page_text: List[str] = []
    ocr_warnings: List[str] = []
    for page_index, page in enumerate(doc, start=1):
        text = page.get_text("text").strip()
        if text:
            page_text.append(f"--- Page {page_index} ---\n{text}")
            continue

        pix = page.get_pixmap(matrix=fitz.Matrix(2, 2), alpha=False)
        ocr_text, warning = _ocr_image_bytes(pix.tobytes("png"))
        if ocr_text:
            page_text.append(f"--- Page {page_index} OCR ---\n{ocr_text}")
        elif warning:
            ocr_warnings.append(f"Page {page_index}: {warning}")

    doc.close()
    extracted = "\n\n".join(page_text).strip()
    warning = " ".join(ocr_warnings)
    method = "PDF text extraction + OCR" if "OCR ---" in extracted else "PDF text extraction"
    return extracted, warning, method


def extract_uploaded_document_text(uploaded_file) -> Tuple[str, str, str]:
    """Extract claim text from uploaded PDFs, images, and text-like files."""
    file_bytes = uploaded_file.getvalue()
    file_type = (uploaded_file.type or "").lower()
    name = uploaded_file.name.lower()

    if file_type == "application/pdf" or name.endswith(".pdf"):
        return extract_pdf_text(file_bytes)

    if file_type.startswith("image/") or name.endswith((".png", ".jpg", ".jpeg", ".tif", ".tiff")):
        text, warning = _ocr_image_bytes(file_bytes)
        return text, warning, "Image OCR"

    if name.endswith((".txt", ".csv", ".tsv")) or file_type.startswith("text/"):
        for encoding in ("utf-8", "utf-16", "latin-1"):
            try:
                return file_bytes.decode(encoding).strip(), "", f"Text decode ({encoding})"
            except UnicodeDecodeError:
                continue
        return "", "Could not decode this text file.", "Text extraction failed"

    return "", "Unsupported document type. Upload a PDF, image, or text file.", "Unsupported"


@st.cache_data(ttl=30, show_spinner=False)
def get_ollama_status() -> Dict[str, Any]:
    """Check local Ollama and pick the best available model."""
    try:
        response = requests.get(OLLAMA_TAGS_URL, timeout=3)
        response.raise_for_status()
        models = response.json().get("models", [])
        names = [m.get("name", "") for m in models if m.get("name")]
        for candidate in [OLLAMA_MODEL, *OLLAMA_FALLBACK_MODELS]:
            if any(name == candidate or name.startswith(f"{candidate}:") for name in names):
                return {
                    "online": True,
                    "model": candidate,
                    "models": names,
                    "using_fallback": candidate != OLLAMA_MODEL,
                    "error": "",
                }
        return {
            "online": True,
            "model": "",
            "models": names,
            "using_fallback": False,
            "error": f"Run `ollama pull {OLLAMA_MODEL}`.",
        }
    except Exception as exc:
        return {
            "online": False,
            "model": "",
            "models": [],
            "using_fallback": False,
            "error": str(exc),
        }


def active_model_label() -> str:
    status = get_ollama_status()
    return status["model"] or OLLAMA_MODEL


def ollama_generate(prompt: str, stream: bool = False) -> Optional[str]:
    """
    Call the local Ollama /api/generate endpoint.
    Returns the model's text response or None on failure.
    """
    status = get_ollama_status()
    if not status["online"]:
        st.error("Ollama is not reachable. Start it with `ollama serve`, then retry.")
        return None
    if not status["model"]:
        available = ", ".join(status["models"]) or "no local models found"
        st.error(
            f"Ollama is running, but `{OLLAMA_MODEL}` is not available. "
            f"Available models: {available}. Run `ollama pull {OLLAMA_MODEL}`."
        )
        return None
    payload = {
        "model" : status["model"],
        "prompt": prompt,
        "stream": stream,
        "options": {"temperature": 0.05, "top_p": 0.9, "num_predict": 1024},
    }
    try:
        response = requests.post(OLLAMA_URL, json=payload, timeout=OLLAMA_TIMEOUT)
        response.raise_for_status()
        return response.json().get("response", "").strip()
    except requests.exceptions.ConnectionError:
        st.error(
            "⚠️ **Ollama not reachable.** "
            "Start the local Ollama server with `ollama serve` and ensure "
            f"`{OLLAMA_MODEL}` is pulled via `ollama pull {OLLAMA_MODEL}`."
        )
    except requests.exceptions.Timeout:
        st.error("⏱️ **LLM request timed out.** The model may be loading — please retry.")
    except requests.exceptions.HTTPError as e:
        st.error(f"🔴 **Ollama HTTP error:** {e}")
    except Exception as e:
        st.error(f"🔴 **Unexpected error calling LLM:** {e}")
    return None


# ── Prompt Template 1: Document Classification ────────────────────────────────
CLASSIFY_PROMPT = """You are an expert insurance document classifier.
Given the OCR text of an insurance-related document, classify it into exactly one
of these categories:
Death Certificate, Hospital Bill, Medical Report, Disability Assessment,
Police Report, Boarding Pass, Insurance Policy, Claim Form, Discharge Summary, Unknown.

Respond with ONLY the category name. No explanation, no punctuation.

DOCUMENT TEXT:
\"\"\"
{text}
\"\"\"

Category:"""


# ── Prompt Template 2: Structured Data Extraction ─────────────────────────────
EXTRACT_PROMPT = """You are a precise insurance data extraction engine.
Extract the following fields from the document text and return a valid JSON object.
If a field is not found, use null. Dates must be in ISO 8601 format (YYYY-MM-DD).
Currency amounts must be numeric (no symbols).

Required fields:
- policy_number       (string)
- insured_name        (string)
- nominee_name        (string)
- nominee_relation    (string)
- accident_date       (date)
- accident_location   (string)
- nature_of_injury    (string)
- diagnosis_codes     (array of strings, ICD-10 codes)
- admission_date      (date)
- discharge_date      (date)
- treatment_costs_inr (number)
- disability_percent  (number or null)
- bank_account_number (string)
- bank_name           (string)
- bank_ifsc           (string)

Return ONLY the JSON object. No markdown, no explanation.

DOCUMENT TEXT:
\"\"\"
{text}
\"\"\"

JSON:"""


# ── Prompt Template 3: Fraud Detection Scoring ───────────────────────────────
FRAUD_PROMPT = """You are a senior insurance fraud analyst specialising in Travel PA claims.
Analyse the document text for anomalies and fraud indicators.

Score the claim on a scale from 0 (no risk) to 100 (definite fraud).
Look for:
- Injury description inconsistent with accident type or mechanism
- Treatment costs unusually high or low for the stated diagnosis
- Dates that are logically impossible (discharge before admission, accident after policy expiry)
- Vague or generic language suggesting fabrication
- Missing or mismatched nominee / bank details
- Diagnosis codes inconsistent with described injury

Respond ONLY with a JSON object in this exact format:
{{
  "fraud_score": <integer 0-100>,
  "risk_level": "<Low|Medium|High|Critical>",
  "flags": ["<flag 1>", "<flag 2>"],
  "summary": "<one-sentence fraud assessment>"
}}

DOCUMENT TEXT:
\"\"\"
{text}
\"\"\"

JSON:"""


VALIDATION_REVIEW_PROMPT = """You are an insurance claims reviewer.
Review the extracted data and deterministic validation results below. Identify any
material inconsistency, missing evidence, or action needed before a claim can proceed.
Do not invent facts that are not present in the supplied data.

Return ONLY valid JSON in this exact format:
{{
  "recommendation": "<Proceed|Manual Review|Request Information>",
  "rationale": "<brief evidence-based explanation>",
  "follow_up_items": ["<specific item>"]
}}

EXTRACTED DATA:
{extracted_json}

VALIDATION RESULTS:
{validation_json}

JSON:"""


def classify_document(text: str) -> str:
    """Use LLM to classify the document type."""
    prompt   = CLASSIFY_PROMPT.format(text=text[:3000])  # token budget
    response = ollama_generate(prompt)
    if response:
        # Sanitise: keep only the first line and match to known types
        first_line = response.strip().split("\n")[0].strip(" .,\"'")
        for doc_type in DOC_TYPES:
            if doc_type.lower() in first_line.lower():
                return doc_type
        return first_line or "Unknown"
    return "Unknown (LLM unavailable)"


def extract_data(text: str) -> Dict[str, Any]:
    """Use LLM to extract structured claim data as a dict."""
    prompt   = EXTRACT_PROMPT.format(text=text[:3000])
    response = ollama_generate(prompt)
    if not response:
        return {}
    # Strip any markdown fences before parsing
    clean = re.sub(r"```(?:json)?|```", "", response).strip()
    try:
        return json.loads(clean)
    except json.JSONDecodeError:
        st.warning("⚠️ LLM returned non-JSON output for extraction; using raw text.")
        return {"raw_response": response}


def score_fraud(text: str) -> Dict[str, Any]:
    """Use LLM to compute a fraud risk score and return flags."""
    prompt   = FRAUD_PROMPT.format(text=text[:3000])
    response = ollama_generate(prompt)
    if not response:
        return {}
    clean = re.sub(r"```(?:json)?|```", "", response).strip()
    try:
        return json.loads(clean)
    except json.JSONDecodeError:
        st.warning("⚠️ LLM returned non-JSON output for fraud scoring; showing raw.")
        return {"raw_response": response}


# ──────────────────────────────────────────────────────────────────────────────
# RULE-BASED VALIDATION ENGINE
# ──────────────────────────────────────────────────────────────────────────────

def review_validation_with_llm(
    extracted: Dict[str, Any], rule_results: List[Dict[str, str]]
) -> Dict[str, Any]:
    """Ask the local model for an evidence-based review of validation findings."""
    prompt = VALIDATION_REVIEW_PROMPT.format(
        extracted_json=json.dumps(extracted, ensure_ascii=False, default=str),
        validation_json=json.dumps(rule_results, ensure_ascii=False, default=str),
    )
    response = ollama_generate(prompt)
    if not response:
        return {}
    clean = response.strip()
    fence = chr(96) * 3
    if clean.startswith(fence):
        clean = clean.split("\n", 1)[-1]
        if clean.endswith(fence):
            clean = clean[:-3]
    try:
        return json.loads(clean.strip())
    except json.JSONDecodeError:
        return {"raw_response": response}


def validate_claim(extracted: Dict[str, Any]) -> List[Dict[str, str]]:
    """
    Run rule-based validations on extracted claim data.
    Returns a list of validation result dicts:
      {"rule": str, "status": "Pass"|"Fail"|"Warning", "message": str}
    """
    results: List[Dict[str, str]] = []

    policy_no  = extracted.get("policy_number")
    acc_date_s = extracted.get("accident_date")
    adm_date_s = extracted.get("admission_date")
    dis_date_s = extracted.get("discharge_date")
    cost       = extracted.get("treatment_costs_inr")
    disability = extracted.get("disability_percent")

    # ── Rule 1: Policy exists ─────────────────────────────────────────────────
    # A PDF can supply a policy number, but a real policy lookup requires a
    # connected policy administration system rather than in-code sample records.
    policy = None
    if not policy_no:
        results.append({"rule": "Policy Number Present", "status": "Fail",
                         "message": "Policy number not found in document."})
    elif not policy:
        results.append({"rule": "Policy Exists", "status": "Warning",
                         "message": f"Policy '{policy_no}' extracted from document; "
                                    "external policy lookup is not connected."})
    else:
        results.append({"rule": "Policy Exists", "status": "Pass",
                         "message": f"Policy {policy_no} verified."})

    # ── Rule 2: Insured name matches ──────────────────────────────────────────
    if policy and extracted.get("insured_name"):
        name_match = (
            extracted["insured_name"].strip().lower()
            == policy["holder"].strip().lower()
        )
        results.append({
            "rule"   : "Insured Name Match",
            "status" : "Pass" if name_match else "Fail",
            "message": "Name matches policy record." if name_match
                       else f"Name mismatch: '{extracted['insured_name']}' vs '{policy['holder']}'.",
        })

    # ── Rule 3: Accident within coverage period ───────────────────────────────
    if policy and acc_date_s:
        try:
            acc_date = datetime.fromisoformat(str(acc_date_s)).date()
            in_period = policy["coverage_start"] <= acc_date <= policy["coverage_end"]
            results.append({
                "rule"   : "Accident Within Coverage Period",
                "status" : "Pass" if in_period else "Fail",
                "message": f"Accident {acc_date} within {policy['coverage_start']} – {policy['coverage_end']}."
                           if in_period
                           else f"Accident on {acc_date} is OUTSIDE coverage period "
                                f"{policy['coverage_start']} – {policy['coverage_end']}.",
            })
        except (ValueError, TypeError):
            results.append({"rule": "Accident Within Coverage Period", "status": "Warning",
                             "message": f"Could not parse accident date: {acc_date_s}"})
    else:
        results.append({"rule": "Accident Date Present", "status": "Warning",
                         "message": "Accident date not extracted."})

    # ── Rule 4: Logical date sequence ─────────────────────────────────────────
    if adm_date_s and dis_date_s:
        try:
            adm = datetime.fromisoformat(str(adm_date_s)).date()
            dis = datetime.fromisoformat(str(dis_date_s)).date()
            if dis >= adm:
                results.append({"rule": "Discharge After Admission", "status": "Pass",
                                 "message": f"Discharged {dis} (admitted {adm})."})
            else:
                results.append({"rule": "Discharge After Admission", "status": "Fail",
                                 "message": f"Discharge ({dis}) precedes admission ({adm}) — impossible dates."})
        except (ValueError, TypeError):
            results.append({"rule": "Discharge After Admission", "status": "Warning",
                             "message": "Could not parse hospitalisation dates."})

    # ── Rule 5: Benefit eligibility ───────────────────────────────────────────
    if policy:
        for cov in policy["coverage_types"]:
            results.append({
                "rule"   : f"Coverage: {cov}",
                "status" : "Pass",
                "message": f"'{cov}' included in policy {policy_no}.",
            })

    # ── Rule 6: Duplicate claim detection ─────────────────────────────────────
    results.append({
        "rule": "Duplicate Claim Check",
        "status": "Warning",
        "message": "Duplicate detection requires a connected claims database.",
    })

    # ── Rule 7: Treatment cost within sum assured ─────────────────────────────
    if policy and cost:
        try:
            cost_f = float(cost)
            if cost_f <= policy["sum_assured"]:
                results.append({"rule": "Cost Within Sum Assured", "status": "Pass",
                                 "message": f"₹{cost_f:,.0f} ≤ Sum Assured ₹{policy['sum_assured']:,.0f}."})
            else:
                results.append({"rule": "Cost Within Sum Assured", "status": "Warning",
                                 "message": f"₹{cost_f:,.0f} exceeds sum assured ₹{policy['sum_assured']:,.0f}."})
        except (ValueError, TypeError):
            pass

    # ── Rule 8: Disability percentage range ──────────────────────────────────
    if disability is not None:
        try:
            pct = float(disability)
            if 0 < pct <= 100:
                results.append({"rule": "Disability % Valid", "status": "Pass",
                                 "message": f"Disability {pct}% is within valid range."})
            else:
                results.append({"rule": "Disability % Valid", "status": "Fail",
                                 "message": f"Disability percentage {pct}% is out of range (1–100)."})
        except (ValueError, TypeError):
            pass

    # ── Rule 9: Bank details present ─────────────────────────────────────────
    has_bank = all(extracted.get(f) for f in ["bank_account_number", "bank_name", "bank_ifsc"])
    results.append({
        "rule"   : "Bank Details Present",
        "status" : "Pass" if has_bank else "Warning",
        "message": "All bank details found." if has_bank else "Incomplete bank details — manual verification required.",
    })

    return results


# ──────────────────────────────────────────────────────────────────────────────
# HELPER UI COMPONENTS
# ──────────────────────────────────────────────────────────────────────────────
def badge_html(status: str) -> str:
    mapping = {
        # Validation rule results
        "Pass"    : "badge-success",
        "Fail"    : "badge-danger",
        "Warning" : "badge-warning",
        # Fraud risk levels
        "Low"     : "badge-success",
        "Medium"  : "badge-warning",
        "High"    : "badge-danger",
        "Critical": "badge-danger",
    }
    # Known document types get the info (blue) style; everything else neutral
    doc_types = {
        "Death Certificate", "Hospital Bill", "Medical Report",
        "Disability Assessment", "Police Report", "Boarding Pass",
        "Insurance Policy", "Claim Form", "Discharge Summary",
    }
    if status in doc_types:
        cls = "badge-info"
    elif status == "Unknown" or status.startswith("Unknown"):
        cls = "badge-neutral"
    else:
        cls = mapping.get(status, "badge-neutral")
    return f'<span class="badge {cls}">{status}</span>'


def metric_card(label: str, value: str, delta: str = "", delta_up: bool = True) -> str:
    delta_cls  = "metric-delta-up" if delta_up else "metric-delta-down"
    delta_html = f'<div class="{delta_cls}">{delta}</div>' if delta else ""
    return f"""
    <div class="card-metric">
        <div class="metric-value">{value}</div>
        <div class="metric-label">{label}</div>
        {delta_html}
    </div>"""


def fraud_bar_html(score: float) -> str:
    if score < 30:
        colour = "#0A7C5C"   # green — low risk
    elif score < 60:
        colour = "#D97706"   # amber — medium risk
    else:
        colour = "#B91C1C"   # red   — high/critical
    return (
        f'<div class="fraud-bar-wrapper">'
        f'<div class="fraud-bar" style="width:{score}%;background:{colour};"></div>'
        f'</div>'
        f'<div class="fraud-score-label">{score:.1f} / 100</div>'
    )


def section_header(title: str) -> None:
    st.markdown(f'<div class="section-header">{title}</div>', unsafe_allow_html=True)


def page_hero(title: str, subtitle: str, kicker: str = "MarvelAI Travel PA") -> None:
    st.markdown(
        f"""
        <div class="page-hero">
            <div class="page-kicker">{kicker}</div>
            <h1>{title}</h1>
            <p>{subtitle}</p>
        </div>
        """,
        unsafe_allow_html=True,
    )


# ──────────────────────────────────────────────────────────────────────────────
# PAGE: DASHBOARD (ANALYTICS)
# ──────────────────────────────────────────────────────────────────────────────
def page_dashboard(df: pd.DataFrame) -> None:
    page_hero(
        "Claims Analytics Dashboard",
        "Real-time operational view of submitted travel PA claims, exposure, risk, and processing velocity.",
        "Portfolio overview",
    )

    # ── Filters ───────────────────────────────────────────────────────────────
    with st.expander("🔍 Filter Claims", expanded=False):
        c1, c2, c3 = st.columns(3)
        with c1:
            sel_cov = st.multiselect("Coverage Type", COVERAGE_TYPES, default=COVERAGE_TYPES)
        with c2:
            sel_geo = st.multiselect("Geography", GEOGRAPHIES, default=GEOGRAPHIES)
        with c3:
            min_d = df["submitted_at"].min().date()
            max_d = df["submitted_at"].max().date()
            date_range = st.date_input("Travel Period", (min_d, max_d))

    # Apply filters
    mask = (
        df["coverage_type"].isin(sel_cov) &
        df["geography"].isin(sel_geo)
    )
    if isinstance(date_range, (list, tuple)) and len(date_range) == 2:
        mask &= (
            (df["submitted_at"].dt.date >= date_range[0]) &
            (df["submitted_at"].dt.date <= date_range[1])
        )
    fdf = df[mask]

    # ── KPI Metric Row ────────────────────────────────────────────────────────
    section_header("Key Performance Indicators")
    m1, m2, m3, m4, m5 = st.columns(5)

    total_claims  = len(fdf)
    total_value   = fdf["claim_amount"].sum()
    approved      = len(fdf[fdf["status"] == "Approved"])
    fraud_flagged = len(fdf[fdf["status"] == "Fraud Flagged"])
    avg_days      = fdf["processing_days"].mean() if total_claims else 0.0

    if total_claims == 0:
        st.warning("No claims match the selected filters. Please broaden your criteria.")
        return

    with m1:
        st.markdown(metric_card("Total Claims", f"{total_claims:,}"), unsafe_allow_html=True)
    with m2:
        st.markdown(metric_card("Total Value", f"₹{total_value/1e6:.2f}M"), unsafe_allow_html=True)
    with m3:
        pct = f"{approved / total_claims * 100:.1f}% ↑"
        st.markdown(metric_card("Approved", f"{approved:,}", pct, True), unsafe_allow_html=True)
    with m4:
        fraud_pct = f"{fraud_flagged / total_claims * 100:.1f}%"
        st.markdown(metric_card("Fraud Flagged", f"{fraud_flagged:,}", fraud_pct, False),
                    unsafe_allow_html=True)
    with m5:
        st.markdown(metric_card("Avg Processing", f"{avg_days:.1f}d"), unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    # ── Charts Row 1 ──────────────────────────────────────────────────────────
    section_header("Claims Distribution")

    # Shared chart constants — defined before column blocks so all charts can use them
    color_map = {
        "Approved"     : "#1F7A4D",
        "Pending"      : "#C98A2E",
        "Rejected"     : "#B2382F",
        "Under Review" : "#2E5EAA",
        "Fraud Flagged": "#7A4F8F",
    }
    chart_layout = dict(
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor ="rgba(0,0,0,0)",
        font=dict(family="Inter, Segoe UI, sans-serif", color="#344054", size=12),
        title_font=dict(size=14, color="#172033", family="Inter, Segoe UI, sans-serif"),
        margin=dict(t=44, b=16, l=16, r=16),
    )

    ch1, ch2 = st.columns(2)

    with ch1:
        status_counts = fdf["status"].value_counts().reset_index()
        status_counts.columns = ["Status", "Count"]

        fig = px.pie(
            status_counts, names="Status", values="Count",
            color="Status", color_discrete_map=color_map,
            title="Claim Status Breakdown",
            hole=0.58,
        )
        fig.update_traces(
            textposition="inside",
            textinfo="percent",
            textfont=dict(size=11, color="#FFFFFF", family="Inter, Segoe UI, sans-serif"),
            insidetextorientation="radial",
        )
        fig.update_layout(
            showlegend=True,
            legend=dict(
                orientation="v",
                x=1.01,
                y=0.5,
                font=dict(color="#344054", size=11),
                bgcolor="rgba(0,0,0,0)",
            ),
            margin=dict(t=44, b=16, l=16, r=120),
            **{k: v for k, v in chart_layout.items() if k != "margin"},
        )
        st.plotly_chart(fig, use_container_width=True)

    with ch2:
        cov_amounts = (
            fdf.groupby("coverage_type")["claim_amount"]
            .sum()
            .reset_index()
            .sort_values("claim_amount", ascending=True)
        )
        fig2 = px.bar(
            cov_amounts,
            x="claim_amount", y="coverage_type",
            orientation="h",
            title="Total Claim Value by Coverage Type",
            labels={"claim_amount": "Amount (₹)", "coverage_type": ""},
            color="claim_amount",
            color_continuous_scale=["#E8F3F1", "#1F6F68"],
        )
        fig2.update_layout(
            coloraxis_showscale=False,
            xaxis=dict(gridcolor="#EEE9E0", tickfont=dict(color="#716B62")),
            yaxis=dict(tickfont=dict(color="#344054")),
            **chart_layout,
        )
        st.plotly_chart(fig2, use_container_width=True)

    # ── Charts Row 2 ──────────────────────────────────────────────────────────
    section_header("Trend & Geography")
    ch3, ch4 = st.columns(2)

    with ch3:
        fdf_copy = fdf.copy()
        fdf_copy["month"] = fdf_copy["submitted_at"].dt.to_period("M").dt.to_timestamp()
        monthly = (
            fdf_copy.groupby(["month", "status"])["claim_id"]
            .count()
            .reset_index()
            .rename(columns={"claim_id": "count"})
        )
        fig3 = px.area(
            monthly, x="month", y="count", color="status",
            title="Monthly Submissions by Status",
            color_discrete_map=color_map,
            labels={"month": "Month", "count": "Claims"},
        )
        fig3.update_layout(
            xaxis=dict(gridcolor="#EEE9E0", tickfont=dict(color="#716B62")),
            yaxis=dict(gridcolor="#EEE9E0", tickfont=dict(color="#716B62")),
            legend=dict(
                title_text="",
                font=dict(color="#344054", size=11),
                bgcolor="rgba(0,0,0,0)",
            ),
            **chart_layout,
        )
        st.plotly_chart(fig3, use_container_width=True)

    with ch4:
        geo_counts = fdf.groupby("geography").agg(
            claims=("claim_id", "count"),
            avg_fraud=("fraud_score", "mean"),
        ).reset_index()
        fig4 = px.scatter(
            geo_counts,
            x="claims", y="avg_fraud",
            size="claims", color="geography",
            text="geography",
            title="Geography: Volume vs Avg Fraud Score",
            labels={"claims": "Claim Count", "avg_fraud": "Avg Fraud Score"},
        )
        fig4.update_traces(
            textposition="top center",
            marker=dict(line=dict(width=1, color="#FFFFFF"), opacity=0.82),
            textfont=dict(size=10, color="#344054"),
        )
        fig4.update_layout(
            showlegend=False,
            xaxis=dict(gridcolor="#EEE9E0", tickfont=dict(color="#716B62")),
            yaxis=dict(gridcolor="#EEE9E0", tickfont=dict(color="#716B62")),
            **chart_layout,
        )
        st.plotly_chart(fig4, use_container_width=True)

    # ── Recent Claims Table ───────────────────────────────────────────────────
    section_header("Recent Claims")
    display_df = fdf.sort_values("submitted_at", ascending=False).head(15).copy()
    display_df["submitted_at"] = display_df["submitted_at"].dt.strftime("%d %b %Y")
    display_df["claim_amount"] = display_df["claim_amount"].apply(lambda x: f"₹{x:,.0f}")
    display_df["fraud_score"]  = display_df["fraud_score"].apply(lambda x: f"{x:.1f}")

    status_badge_map = {
        "Approved"     : ("#E9F6EE", "#1F7A4D"),
        "Pending"      : ("#FFF8E8", "#A86512"),
        "Rejected"     : ("#FCEDEA", "#B2382F"),
        "Under Review" : ("#EAF1FB", "#2E5EAA"),
        "Fraud Flagged": ("#F3EEF9", "#7A4F8F"),
    }

    col_labels = ["Claim ID", "Submitted", "Coverage", "Geography",
                  "Status", "Amount", "Fraud Score", "Days"]
    col_keys   = ["claim_id", "submitted_at", "coverage_type", "geography",
                  "status", "claim_amount", "fraud_score", "processing_days"]

    header_cells = "".join(
        f"""<th style="background:#F0EAE0;color:#172033;font-size:0.7rem;font-weight:700;
            text-transform:uppercase;letter-spacing:0.08em;padding:0.55rem 0.85rem;
            border-bottom:2px solid #D6CEC0;white-space:nowrap;">{lbl}</th>"""
        for lbl in col_labels
    )

    rows_html = ""
    for _, row in display_df.iterrows():
        cells = ""
        for key in col_keys:
            val = row[key]
            if key == "status":
                bg, fg = status_badge_map.get(str(val), ("#F0EAE0", "#344054"))
                cell_html = (
                    f"""<span style="display:inline-block;padding:0.15rem 0.55rem;"
                    f"border-radius:999px;background:{bg};color:{fg};"
                    f"font-size:0.68rem;font-weight:700;">{val}</span>"""
                )
                cells += f"""<td style="padding:0.55rem 0.85rem;">{cell_html}</td>"""
            elif key == "fraud_score":
                score_f = float(val)
                color = "#B2382F" if score_f >= 60 else ("#A86512" if score_f >= 30 else "#1F7A4D")
                cells += (
                    f"""<td style="padding:0.55rem 0.85rem;color:{color};"
                    f"font-weight:600;font-size:0.83rem;">{val}</td>"""
                )
            else:
                cells += f"""<td style="padding:0.55rem 0.85rem;color:#344054;font-size:0.83rem;">{val}</td>"""
        rows_html += f"<tr style='border-bottom:1px solid #EEE9E0;'>{cells}</tr>"

    table_html = f"""
    <div style="overflow-x:auto;border-radius:8px;border:1px solid #D6CEC0;
                box-shadow:0 8px 24px rgba(23,32,51,0.045);">
      <table style="width:100%;border-collapse:collapse;background:#FFFFFF;
                    font-family:Inter,Segoe UI,sans-serif;">
        <thead><tr>{header_cells}</tr></thead>
        <tbody>{rows_html}</tbody>
      </table>
    </div>"""
    st.markdown(table_html, unsafe_allow_html=True)


# ──────────────────────────────────────────────────────────────────────────────
# PAGE: CLAIM INGESTION
# ──────────────────────────────────────────────────────────────────────────────
def page_ingestion() -> None:
    page_hero(
        "Claim Document Ingestion",
        "Upload claim files and extract real document text for downstream LLM classification, validation, and fraud review.",
        "Intake desk",
    )

    uploaded_files = st.file_uploader(
        "Drop claim documents here",
        type=["pdf", "png", "jpg", "jpeg", "tif", "tiff", "txt", "csv", "tsv"],
        accept_multiple_files=True,
        help="Accepted formats: PDF, PNG, JPG, TIFF — max 10 MB per file",
    )

    if not uploaded_files:
        st.info("👆 Upload one or more documents to begin.")
        return

    section_header(f"{len(uploaded_files)} Document(s) Uploaded")

    for file in uploaded_files:
        with st.container():
            st.markdown('<div class="card">', unsafe_allow_html=True)
            c1, c2 = st.columns([3, 1])
            with c1:
                st.markdown(
                    f"**{file.name}**&nbsp; "
                    f"<span style='color:#5A6E89;font-size:0.8rem;font-weight:400;'>"
                    f"{file.size / 1024:.1f} KB &nbsp;·&nbsp; {file.type}</span>",
                    unsafe_allow_html=True,
                )
            with c2:
                if file.type.startswith("image/"):
                    st.image(file, use_container_width=True)
                else:
                    st.markdown("📄 PDF document")

            ocr_key  = f"ocr_{file.name}"
            meta_key = f"meta_{file.name}"

            file_hash = hashlib.sha256(file.getvalue()).hexdigest()
            # Re-extract if the file contents change, even when the file name is reused.
            if (
                ocr_key not in st.session_state
                or st.session_state.get(meta_key, {}).get("file_hash") != file_hash
            ):
                with st.spinner(f"Extracting text from {file.name}…"):
                    extracted_text, warning, method = extract_uploaded_document_text(file)

                if not extracted_text.strip():
                    extracted_text = ""
                    method = "Text extraction failed"
                    if warning:
                        st.warning(f"⚠️ Extraction note: {warning}")
                    st.error(
                        "No readable text was found in this document. "
                        "For scanned PDFs, install and configure Tesseract OCR."
                    )
                else:
                    if warning:
                        st.warning(f"⚠️ Extraction note: {warning}")
                    st.success(f"✅ Text extracted via **{method}** ({len(extracted_text):,} chars)")

                st.session_state[ocr_key]  = extracted_text
                st.session_state[meta_key] = {
                    "filename"   : file.name,
                    "size_kb"    : round(file.size / 1024, 1),
                    "upload_ts"  : datetime.now().isoformat(),
                    "claim_ref"  : f"CLM-{uuid.uuid4().hex[:8].upper()}",
                    "ocr_method" : method,
                    "file_hash"  : file_hash,
                    "extraction_warning": warning,
                }
                for prefix in ("cls_", "ext_", "val_", "fraud_", "ai_val_"):
                    st.session_state.pop(f"{prefix}{file.name}", None)

            meta = st.session_state[meta_key]
            st.markdown(
                f"**Claim Ref:** `{meta['claim_ref']}`  &nbsp;|&nbsp;  "
                f"**Uploaded:** {meta['upload_ts'][:19].replace('T', ' ')}  &nbsp;|&nbsp;  "
                f"**Method:** {meta.get('ocr_method', 'N/A')}",
                unsafe_allow_html=True,
            )

            with st.expander("🔍 View Extracted Text"):
                st.code(st.session_state[ocr_key], language="text")

            st.markdown("</div>", unsafe_allow_html=True)

    if uploaded_files:
        st.success(
            f"✅ {len(uploaded_files)} document(s) ingested. "
            "Proceed to **Processing & Extraction** in the sidebar."
        )


# ──────────────────────────────────────────────────────────────────────────────
# PAGE: PROCESSING & EXTRACTION
# ──────────────────────────────────────────────────────────────────────────────
def _render_json_pre(data: dict) -> None:
    """Render a dict as a styled, light-themed <pre> block."""
    import json as _json
    json_str = _json.dumps(data, indent=2, ensure_ascii=False)
    escaped  = json_str.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
    st.markdown(
        f'''<pre style="background:#F4F0E8;border:1px solid #D6CEC0;border-radius:8px;
        padding:1rem 1.1rem;font-family:'JetBrains Mono','Fira Code',Consolas,monospace;
        font-size:0.78rem;color:#344054;line-height:1.75;overflow-x:auto;
        white-space:pre-wrap;word-break:break-word;">{escaped}</pre>''',
        unsafe_allow_html=True,
    )


def page_processing() -> None:
    page_hero(
        "Processing & Data Extraction",
        f"Classify each document and extract structured fields using local Ollama model `{active_model_label()}`.",
        "AI document operations",
    )

    ocr_keys = [k for k in st.session_state if k.startswith("ocr_")]
    if not ocr_keys:
        st.warning("No documents ingested yet. Go to **Claim Ingestion** first.")
        return

    # ── "Process All" shortcut ────────────────────────────────────────────────
    if st.button("⚡ Process All Documents", type="primary", use_container_width=False):
        for ocr_key in ocr_keys:
            filename = ocr_key.replace("ocr_", "")
            text = st.session_state[ocr_key]
            if not text.strip():
                continue
            with st.spinner(f"Classifying {filename}…"):
                st.session_state[f"cls_{filename}"] = classify_document(text)
            with st.spinner(f"Extracting fields from {filename}…"):
                st.session_state[f"ext_{filename}"] = extract_data(text)
        st.success("All documents processed.")
        st.rerun()

    st.markdown("<br>", unsafe_allow_html=True)

    for ocr_key in ocr_keys:
        filename = ocr_key.replace("ocr_", "")
        meta_key = f"meta_{filename}"
        meta     = st.session_state.get(meta_key, {})
        text     = st.session_state[ocr_key]
        cls_key  = f"cls_{filename}"
        ext_key  = f"ext_{filename}"

        if not text.strip():
            st.error(
                f"{filename}: no text could be extracted. Upload a text-based PDF "
                "or configure OCR for scanned documents."
            )
            continue

        already_done = cls_key in st.session_state and ext_key in st.session_state

        st.markdown('<div class="card">', unsafe_allow_html=True)

        hdr_col, btn_col = st.columns([4, 1])
        with hdr_col:
            st.markdown(f"### 📄 {filename}")
            st.markdown(
                f"Claim Ref: `{meta.get('claim_ref', 'N/A')}`  &nbsp;·&nbsp;  "
                f"{len(text):,} characters of extracted text",
                unsafe_allow_html=True,
            )
        with btn_col:
            run_label = "🔄 Re-process" if already_done else "▶ Process"
            if st.button(run_label, key=f"btn_proc_{filename}"):
                with st.spinner("Classifying…"):
                    st.session_state[cls_key] = classify_document(text)
                with st.spinner("Extracting structured fields…"):
                    st.session_state[ext_key] = extract_data(text)
                st.rerun()

        # Show results only once both steps have run
        if cls_key in st.session_state:
            section_header("Step 1 · Document Classification")
            st.markdown(
                f"**Detected Type:** {badge_html(st.session_state[cls_key])}",
                unsafe_allow_html=True,
            )

        if ext_key in st.session_state:
            st.markdown("<br>", unsafe_allow_html=True)
            section_header("Step 2 · Structured Data Extraction")
            extracted = st.session_state[ext_key]
            if "raw_response" in extracted:
                st.warning("LLM returned non-JSON; showing raw response:")
                st.code(extracted["raw_response"], language="text")
            else:
                # Render each field as a readable key-value grid
                field_labels = {
                    "policy_number"       : "Policy Number",
                    "insured_name"        : "Insured Name",
                    "nominee_name"        : "Nominee Name",
                    "nominee_relation"    : "Nominee Relation",
                    "accident_date"       : "Accident Date",
                    "accident_location"   : "Accident Location",
                    "nature_of_injury"    : "Nature of Injury",
                    "diagnosis_codes"     : "Diagnosis Codes (ICD-10)",
                    "admission_date"      : "Admission Date",
                    "discharge_date"      : "Discharge Date",
                    "treatment_costs_inr" : "Treatment Costs (₹)",
                    "disability_percent"  : "Disability %",
                    "bank_account_number" : "Bank Account",
                    "bank_name"           : "Bank Name",
                    "bank_ifsc"           : "IFSC Code",
                }
                rows = []
                for key, label in field_labels.items():
                    val = extracted.get(key)
                    if isinstance(val, list):
                        val = ", ".join(str(v) for v in val) if val else "—"
                    elif val is None or val == "":
                        val = "—"
                    rows.append((label, str(val)))

                # Render as a two-column HTML table
                cells_html = "".join(
                    f'''<tr>
                      <td style="padding:0.45rem 0.9rem;font-size:0.78rem;font-weight:600;
                                 color:#5A6E89;white-space:nowrap;border-bottom:1px solid #EEE9E0;
                                 width:38%;">{label}</td>
                      <td style="padding:0.45rem 0.9rem;font-size:0.82rem;color:#172033;
                                 border-bottom:1px solid #EEE9E0;">{value}</td>
                    </tr>'''
                    for label, value in rows
                )
                st.markdown(
                    f'''<div style="overflow-x:auto;border-radius:8px;border:1px solid #D6CEC0;">
                      <table style="width:100%;border-collapse:collapse;background:#FFFFFF;
                                    font-family:Inter,Segoe UI,sans-serif;">
                        {cells_html}
                      </table>
                    </div>''',
                    unsafe_allow_html=True,
                )
                with st.expander("📋 View raw JSON"):
                    _render_json_pre(extracted)

        if not (cls_key in st.session_state or ext_key in st.session_state):
            st.caption("Click **▶ Process** to classify this document and extract fields.")

        st.markdown("</div>", unsafe_allow_html=True)
        st.markdown("<br>", unsafe_allow_html=True)


# ──────────────────────────────────────────────────────────────────────────────
# PAGE: VALIDATION
# ──────────────────────────────────────────────────────────────────────────────
def page_validation() -> None:
    page_hero(
        "Claim Validation",
        "Run policy, chronology, coverage, and payout checks against extracted claim data.",
        "Rules engine",
    )

    ext_keys = [k for k in st.session_state if k.startswith("ext_")]
    if not ext_keys:
        st.warning(
            "No extracted data found. Complete **Processing & Extraction** first."
        )
        return

    for ext_key in ext_keys:
        filename = ext_key.replace("ext_", "")
        extracted: Dict = st.session_state[ext_key]
        meta: Dict      = st.session_state.get(f"meta_{filename}", {})

        st.markdown('<div class="card">', unsafe_allow_html=True)
        st.markdown(f"### 📄 {filename}  —  `{meta.get('claim_ref', '')}`")

        val_key = f"val_{filename}"
        ai_key = f"ai_val_{filename}"
        run_btn = st.button("▶ Run Validation Rules", key=f"btn_val_{filename}")
        if run_btn:
            results = validate_claim(extracted)
            st.session_state[val_key] = results
            with st.spinner(f"Reviewing validation with {active_model_label()}…"):
                st.session_state[ai_key] = review_validation_with_llm(extracted, results)

        if val_key in st.session_state:
            results = st.session_state[val_key]

            llm_review = st.session_state.get(ai_key, {})
            if llm_review:
                section_header(f"Local LLM Review ({active_model_label()})")
                if "raw_response" in llm_review:
                    st.code(llm_review["raw_response"], language="text")
                else:
                    recommendation = llm_review.get("recommendation", "Manual Review")
                    st.markdown(f"**Recommendation:** {recommendation}")
                    if llm_review.get("rationale"):
                        st.caption(llm_review["rationale"])
                    for item in llm_review.get("follow_up_items", []):
                        st.markdown(f"- {item}")
                st.divider()

            # Summary counts
            passes   = sum(1 for r in results if r["status"] == "Pass")
            fails    = sum(1 for r in results if r["status"] == "Fail")
            warnings = sum(1 for r in results if r["status"] == "Warning")

            mc1, mc2, mc3 = st.columns(3)
            with mc1:
                st.markdown(
                    metric_card("Passed", str(passes), "✔", True),
                    unsafe_allow_html=True,
                )
            with mc2:
                st.markdown(
                    metric_card("Failed", str(fails), "✖", False),
                    unsafe_allow_html=True,
                )
            with mc3:
                st.markdown(
                    metric_card("Warnings", str(warnings), "⚠", False),
                    unsafe_allow_html=True,
                )

            st.markdown("<br>", unsafe_allow_html=True)
            section_header("Rule-by-Rule Results")

            for r in results:
                badge  = badge_html(r["status"])
                icon   = "✅" if r["status"] == "Pass" else ("❌" if r["status"] == "Fail" else "⚠️")
                st.markdown(
                    f"{icon} &nbsp; **{r['rule']}** &nbsp; {badge}  \n"
                    f"<small style='color:#5A6E89;'>{r['message']}</small>",
                    unsafe_allow_html=True,
                )
                st.divider()

            # Overall recommendation
            if fails == 0 and warnings == 0:
                st.success("🎉 All validation rules passed. Claim is eligible for processing.")
            elif fails > 0:
                st.error(f"🚫 {fails} rule(s) failed. Claim requires manual review before approval.")
            else:
                st.warning(f"⚠️ {warnings} warning(s) found. Secondary verification recommended.")

        st.markdown("</div>", unsafe_allow_html=True)
        st.markdown("<br>", unsafe_allow_html=True)


# ──────────────────────────────────────────────────────────────────────────────
# PAGE: AI FRAUD SCORING
# ──────────────────────────────────────────────────────────────────────────────
def page_fraud_scoring() -> None:
    page_hero(
        "AI Fraud Detection & Scoring",
        f"Use `{active_model_label()}` through Ollama to assess anomaly signals, fraud flags, and escalation priority.",
        "Local LLM risk review",
    )

    ocr_keys = [k for k in st.session_state if k.startswith("ocr_")]
    if not ocr_keys:
        st.warning("No documents ingested. Go to **Claim Ingestion** first.")
        return

    for ocr_key in ocr_keys:
        filename = ocr_key.replace("ocr_", "")
        text     = st.session_state[ocr_key]
        meta     = st.session_state.get(f"meta_{filename}", {})

        st.markdown('<div class="card">', unsafe_allow_html=True)
        st.markdown(f"### 📄 {filename}  —  `{meta.get('claim_ref', '')}`")

        fraud_key = f"fraud_{filename}"
        if st.button("🔍 Run Fraud Analysis", key=f"btn_fraud_{filename}"):
            with st.spinner("LLM analysing for fraud signals…"):
                result = score_fraud(text)
                st.session_state[fraud_key] = result

        if fraud_key in st.session_state:
            result = st.session_state[fraud_key]

            if "raw_response" in result:
                st.code(result["raw_response"], language="json")
            else:
                score      = result.get("fraud_score", 0)
                risk_level = result.get("risk_level", "Unknown")
                flags      = result.get("flags", [])
                summary    = result.get("summary", "")

                # Score gauge
                fc1, fc2 = st.columns([1, 2])
                with fc1:
                    section_header("Fraud Score")
                    score_colour = (
                        "#B91C1C" if score >= 60
                        else "#7A4F00" if score >= 30
                        else "#0A7C5C"
                    )
                    st.markdown(
                        f'<div style="font-size:2.8rem;font-weight:700;'
                        f'color:{score_colour};letter-spacing:-0.03em;">'
                        f'{score}</div>',
                        unsafe_allow_html=True,
                    )
                    st.markdown(fraud_bar_html(score), unsafe_allow_html=True)
                    st.markdown(
                        f"**Risk Level:** {badge_html(risk_level)}",
                        unsafe_allow_html=True,
                    )

                with fc2:
                    section_header("Fraud Flags Detected")
                    if flags:
                        for flag in flags:
                            st.markdown(f"🚩 {flag}")
                    else:
                        st.markdown("✅ No specific flags raised.")

                    if summary:
                        st.markdown("---")
                        st.markdown(f"**Assessment:** _{summary}_")

                # Recommendation
                st.markdown("<br>", unsafe_allow_html=True)
                if score < 30:
                    st.success("✅ Low fraud risk. Proceed with standard processing.")
                elif score < 60:
                    st.warning("⚠️ Medium fraud risk. Additional verification recommended.")
                else:
                    st.error("🚨 High / Critical fraud risk. Escalate to Special Investigations Unit (SIU).")

        else:
            st.caption("Click **Run Fraud Analysis** to score this document.")

        st.markdown("</div>", unsafe_allow_html=True)
        st.markdown("<br>", unsafe_allow_html=True)




# ──────────────────────────────────────────────────────────────────────────────
# PAGE: CLAIM REPORT
# ──────────────────────────────────────────────────────────────────────────────
def _excel_value(value: Any) -> Any:
    """Return a spreadsheet-safe representation of an extracted value."""
    if isinstance(value, list):
        return ", ".join(str(item) for item in value)
    if isinstance(value, (dict, tuple)):
        return json.dumps(value, ensure_ascii=False)
    return "" if value is None else value


def _report_excel_bytes(
    filename: str,
    meta: dict,
    doc_type: str,
    extracted: dict,
    val_results: list,
    fraud: dict,
) -> bytes:
    """Create a formatted Excel workbook containing the claim report data."""
    try:
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise RuntimeError(
            "Excel export requires openpyxl. Install it with: pip install openpyxl"
        ) from exc

    claim_ref = meta.get("claim_ref", "N/A")
    generated_at = datetime.now().strftime("%d %b %Y %H:%M")
    fail_count = sum(rule.get("status") == "Fail" for rule in val_results)
    warning_count = sum(rule.get("status") == "Warning" for rule in val_results)
    verdict = (
        "Eligible for processing" if fail_count == 0 and warning_count == 0
        else "Requires manual review" if fail_count else "Secondary verification recommended"
    )

    summary_df = pd.DataFrame([
        ("Claim reference", claim_ref),
        ("Source document", filename),
        ("Document type", doc_type),
        ("Uploaded", meta.get("upload_ts", "")),
        ("Extraction method", meta.get("ocr_method", meta.get("extraction_method", "N/A"))),
        ("Generated", generated_at),
        ("Validation verdict", verdict),
        ("Fraud score", fraud.get("fraud_score", "Not run")),
        ("Fraud risk level", fraud.get("risk_level", "Not run")),
    ], columns=["Report item", "Value"])
    extracted_df = pd.DataFrame([
        {"Field": key.replace("_", " ").title(), "Value": _excel_value(value)}
        for key, value in extracted.items()
    ])
    validation_df = pd.DataFrame(val_results or [], columns=["rule", "status", "message"])
    validation_df = validation_df.rename(columns={
        "rule": "Rule", "status": "Status", "message": "Details"
    })
    fraud_df = pd.DataFrame([
        ("Fraud score", _excel_value(fraud.get("fraud_score", "Not run"))),
        ("Risk level", _excel_value(fraud.get("risk_level", "Not run"))),
        ("Flags", _excel_value(fraud.get("flags", []))),
        ("Assessment", _excel_value(fraud.get("summary", ""))),
    ], columns=["Assessment item", "Value"])

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="Summary", index=False)
        extracted_df.to_excel(writer, sheet_name="Extracted Data", index=False)
        validation_df.to_excel(writer, sheet_name="Validation", index=False)
        fraud_df.to_excel(writer, sheet_name="Fraud Review", index=False)

        header_fill = PatternFill("solid", fgColor="1F6F68")
        header_font = Font(color="FFFFFF", bold=True)
        for worksheet in writer.book.worksheets:
            worksheet.freeze_panes = "A2"
            worksheet.auto_filter.ref = worksheet.dimensions
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(vertical="center")
            for column_cells in worksheet.columns:
                width = min(
                    max(len(str(cell.value or "")) for cell in column_cells) + 2,
                    60,
                )
                worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = width
            for row in worksheet.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = Alignment(vertical="top", wrap_text=True)

    return output.getvalue()


def _report_html(filename: str, meta: dict, doc_type: str,
                 extracted: dict, val_results: list, fraud: dict) -> str:
    """Build a self-contained HTML report string for one claim."""
    import html as _html

    claim_ref  = meta.get("claim_ref", "N/A")
    upload_ts  = meta.get("upload_ts", "")[:19].replace("T", " ")
    ocr_method = meta.get("ocr_method", "N/A")

    # ── Extracted fields table ────────────────────────────────────────────────
    field_labels = {
        "policy_number"       : "Policy Number",
        "insured_name"        : "Insured Name",
        "nominee_name"        : "Nominee Name",
        "nominee_relation"    : "Nominee Relation",
        "accident_date"       : "Accident Date",
        "accident_location"   : "Accident Location",
        "nature_of_injury"    : "Nature of Injury",
        "diagnosis_codes"     : "Diagnosis Codes",
        "admission_date"      : "Admission Date",
        "discharge_date"      : "Discharge Date",
        "treatment_costs_inr" : "Treatment Costs (₹)",
        "disability_percent"  : "Disability %",
        "bank_account_number" : "Bank Account",
        "bank_name"           : "Bank",
        "bank_ifsc"           : "IFSC Code",
    }
    ext_rows = ""
    for key, label in field_labels.items():
        val = extracted.get(key)
        if isinstance(val, list):
            val = ", ".join(str(v) for v in val) if val else "—"
        elif val is None or val == "":
            val = "—"
        ext_rows += f"""
        <tr>
          <td style="padding:6px 12px;font-weight:600;color:#5A6E89;
                     width:38%;border-bottom:1px solid #EEE9E0;">{label}</td>
          <td style="padding:6px 12px;color:#172033;
                     border-bottom:1px solid #EEE9E0;">{_html.escape(str(val))}</td>
        </tr>"""

    # ── Validation results table ───────────────────────────────────────────────
    status_colours = {"Pass": "#1F7A4D", "Fail": "#B2382F", "Warning": "#A86512"}
    status_bg      = {"Pass": "#E9F6EE", "Fail": "#FCEDEA", "Warning": "#FFF8E8"}
    val_rows = ""
    passes = fails = warnings = 0
    for r in val_results:
        s = r["status"]
        passes   += s == "Pass"
        fails    += s == "Fail"
        warnings += s == "Warning"
        fg = status_colours.get(s, "#344054")
        bg = status_bg.get(s, "#F4F0E8")
        val_rows += f"""
        <tr>
          <td style="padding:6px 12px;border-bottom:1px solid #EEE9E0;color:#172033;">
              {_html.escape(r["rule"])}</td>
          <td style="padding:6px 12px;border-bottom:1px solid #EEE9E0;">
              <span style="background:{bg};color:{fg};padding:2px 10px;border-radius:999px;
                           font-size:0.72rem;font-weight:700;">{s}</span></td>
          <td style="padding:6px 12px;border-bottom:1px solid #EEE9E0;color:#5A6E89;
                     font-size:0.8rem;">{_html.escape(r["message"])}</td>
        </tr>"""

    val_summary = (
        f"{passes} passed &nbsp;·&nbsp; {fails} failed &nbsp;·&nbsp; {warnings} warnings"
    )
    overall_verdict = (
        "✅ ELIGIBLE FOR PROCESSING" if fails == 0 and warnings == 0
        else ("🚫 REQUIRES MANUAL REVIEW" if fails > 0
              else "⚠️ SECONDARY VERIFICATION RECOMMENDED")
    )
    verdict_colour = "#1F7A4D" if fails == 0 and warnings == 0 else (
        "#B2382F" if fails > 0 else "#A86512"
    )

    # ── Fraud section ─────────────────────────────────────────────────────────
    fraud_score = fraud.get("fraud_score", "N/A")
    risk_level  = fraud.get("risk_level", "N/A")
    flags       = fraud.get("flags", [])
    fraud_summary = fraud.get("summary", "N/A")
    fraud_bar_pct = fraud_score if isinstance(fraud_score, (int, float)) else 0
    fraud_bar_col = (
        "#B91C1C" if fraud_bar_pct >= 60
        else "#D97706" if fraud_bar_pct >= 30
        else "#0A7C5C"
    )
    flags_html = "".join(
        f'<li style="margin:4px 0;color:#344054;">🚩 {_html.escape(f)}</li>'
        for f in flags
    ) or '<li style="color:#1F7A4D;">✅ No specific flags raised.</li>'

    risk_colours = {
        "Low": ("#E9F6EE", "#1F7A4D"),
        "Medium": ("#FFF8E8", "#A86512"),
        "High": ("#FCEDEA", "#B2382F"),
        "Critical": ("#FCEDEA", "#B2382F"),
    }
    rl_bg, rl_fg = risk_colours.get(risk_level, ("#F4F0E8", "#344054"))

    recommendation = (
        "Proceed with standard processing." if fraud_bar_pct < 30
        else ("Additional verification recommended." if fraud_bar_pct < 60
              else "ESCALATE TO SPECIAL INVESTIGATIONS UNIT (SIU).")
    )

    # ── Assemble full HTML ────────────────────────────────────────────────────
    return f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<title>Claim Report — {claim_ref}</title>
<style>
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{ font-family: Inter, Segoe UI, Arial, sans-serif; font-size: 13px;
          color: #172033; background: #FFFFFF; padding: 32px 40px; }}
  h1 {{ font-size: 1.55rem; font-weight: 800; color: #172033; }}
  h2 {{ font-size: 0.95rem; font-weight: 700; color: #172033;
        text-transform: uppercase; letter-spacing: 0.08em;
        border-bottom: 2px solid #C98A2E; padding-bottom: 6px;
        margin: 28px 0 14px; }}
  table {{ width: 100%; border-collapse: collapse; background: #FFFFFF; }}
  .badge {{ display: inline-block; padding: 2px 10px; border-radius: 999px;
            font-size: 0.7rem; font-weight: 700; }}
  .header-bar {{ display: flex; justify-content: space-between;
                 align-items: flex-start; margin-bottom: 24px; }}
  .header-meta {{ font-size: 0.8rem; color: #5A6E89; line-height: 1.8; }}
  .fraud-bar-outer {{ background: #EEE9E0; border-radius: 999px;
                      height: 12px; width: 100%; margin: 6px 0; }}
  .fraud-bar-inner {{ height: 12px; border-radius: 999px;
                      background: {fraud_bar_col};
                      width: {fraud_bar_pct}%; }}
  .verdict {{ font-size: 0.92rem; font-weight: 700; color: {verdict_colour};
              background: {"#E9F6EE" if verdict_colour == "#1F7A4D" else ("#FCEDEA" if verdict_colour == "#B2382F" else "#FFF8E8")};
              padding: 10px 16px; border-radius: 8px; margin-top: 12px; }}
  .footer {{ margin-top: 36px; font-size: 0.7rem; color: #A0AEC0;
             border-top: 1px solid #EEE9E0; padding-top: 10px; }}
  @media print {{ body {{ padding: 12px 16px; }} }}
</style>
</head>
<body>

<div class="header-bar">
  <div>
    <div style="font-size:0.72rem;font-weight:700;color:#C98A2E;
                letter-spacing:0.13em;text-transform:uppercase;margin-bottom:4px;">
        MarvelAI · Travel PA Claims Portal
    </div>
    <h1>Claim Analysis Report</h1>
    <div style="margin-top:8px;font-size:0.88rem;color:#5A6E89;">
        {_html.escape(filename)} &nbsp;·&nbsp;
        <span style="background:#EAF1FB;color:#2E5EAA;padding:2px 9px;border-radius:999px;
                     font-size:0.72rem;font-weight:700;">{_html.escape(doc_type)}</span>
    </div>
  </div>
  <div class="header-meta" style="text-align:right;">
    <strong>Claim Ref</strong>&nbsp; {claim_ref}<br>
    <strong>Uploaded</strong>&nbsp; {upload_ts}<br>
    <strong>Extraction</strong>&nbsp; {_html.escape(ocr_method)}<br>
    <strong>Generated</strong>&nbsp; {datetime.now().strftime("%d %b %Y %H:%M")}
  </div>
</div>

<!-- ── Extracted Claim Data ────────────────────────────────────────────── -->
<h2>Extracted Claim Data</h2>
<table>
  <tbody>{ext_rows}</tbody>
</table>

<!-- ── Validation Results ─────────────────────────────────────────────── -->
<h2>Validation Results &nbsp;<span style="font-size:0.78rem;font-weight:400;
    color:#5A6E89;text-transform:none;letter-spacing:0;">{val_summary}</span></h2>
<table>
  <thead>
    <tr style="background:#F0EAE0;">
      <th style="padding:7px 12px;text-align:left;font-size:0.72rem;font-weight:700;
                 text-transform:uppercase;color:#5A6E89;letter-spacing:0.06em;width:35%;">Rule</th>
      <th style="padding:7px 12px;text-align:left;font-size:0.72rem;font-weight:700;
                 text-transform:uppercase;color:#5A6E89;letter-spacing:0.06em;width:12%;">Status</th>
      <th style="padding:7px 12px;text-align:left;font-size:0.72rem;font-weight:700;
                 text-transform:uppercase;color:#5A6E89;letter-spacing:0.06em;">Details</th>
    </tr>
  </thead>
  <tbody>{val_rows}</tbody>
</table>
<div class="verdict">{overall_verdict}</div>

<!-- ── Fraud Scoring ──────────────────────────────────────────────────── -->
<h2>Fraud Risk Assessment</h2>
<table>
  <tr>
    <td style="padding:8px 12px;width:50%;vertical-align:top;">
      <div style="font-size:2.4rem;font-weight:800;color:{fraud_bar_col};
                  letter-spacing:-0.03em;">{fraud_score}</div>
      <div style="font-size:0.72rem;color:#5A6E89;margin-bottom:4px;">Fraud Score (0–100)</div>
      <div class="fraud-bar-outer"><div class="fraud-bar-inner"></div></div>
      <div style="margin-top:8px;">
        Risk Level: &nbsp;
        <span class="badge" style="background:{rl_bg};color:{rl_fg};">{risk_level}</span>
      </div>
    </td>
    <td style="padding:8px 12px;vertical-align:top;">
      <strong style="font-size:0.78rem;color:#5A6E89;text-transform:uppercase;
                     letter-spacing:0.07em;">Flags Detected</strong>
      <ul style="margin-top:8px;padding-left:18px;">{flags_html}</ul>
      <div style="margin-top:12px;font-size:0.82rem;color:#344054;">
        <strong>Assessment:</strong> {_html.escape(fraud_summary)}
      </div>
      <div style="margin-top:8px;font-size:0.82rem;font-weight:700;color:{fraud_bar_col};">
        Recommendation: {recommendation}
      </div>
    </td>
  </tr>
</table>

<div class="footer">
  MarvelAI Travel PA Claims Portal · Auto-generated report · {datetime.now().strftime("%d %b %Y %H:%M UTC")}
</div>

</body>
</html>"""


def page_report() -> None:
    page_hero(
        "Claim Analysis Report",
        "Compile extracted data, validation results, and fraud scores into a downloadable report for each processed claim.",
        "Report generation",
    )

    ocr_keys = [k for k in st.session_state if k.startswith("ocr_")]
    if not ocr_keys:
        st.warning("No documents in session. Go to **Claim Ingestion** first.")
        return

    # Check which documents have been fully processed
    ready    = []
    not_done = []
    for ocr_key in ocr_keys:
        fname = ocr_key.replace("ocr_", "")
        has_ext = f"ext_{fname}" in st.session_state
        has_val = f"val_{fname}" in st.session_state
        has_fraud = f"fraud_{fname}" in st.session_state
        if has_ext:
            ready.append(fname)
        else:
            not_done.append(fname)

    if not_done:
        st.info(
            f"⚠️ {len(not_done)} document(s) haven't been processed yet: "
            + ", ".join(f"**{f}**" for f in not_done)
            + ". Visit **Processing & Extraction** (and optionally Validation & "
            "Fraud Scoring) to enrich the report."
        )

    if not ready:
        st.warning("No documents have been processed. Run **Processing & Extraction** first.")
        return

    section_header(f"{len(ready)} Document Report(s) Ready")

    for filename in ready:
        meta      = st.session_state.get(f"meta_{filename}", {})
        doc_type  = st.session_state.get(f"cls_{filename}", "Unknown")
        extracted = st.session_state.get(f"ext_{filename}", {})
        val_results = st.session_state.get(f"val_{filename}", [])
        fraud     = st.session_state.get(f"fraud_{filename}", {})

        claim_ref = meta.get("claim_ref", "N/A")

        # Run validation automatically if not done yet
        if not val_results and extracted:
            val_results = validate_claim(extracted)
            st.session_state[f"val_{filename}"] = val_results

        st.markdown('<div class="card">', unsafe_allow_html=True)
        hdr, dl_col = st.columns([4, 1])
        with hdr:
            st.markdown(f"### 📄 {filename}  —  `{claim_ref}`")
            status_parts = []
            status_parts.append(f"✅ Extraction")
            if val_results:
                fails = sum(1 for r in val_results if r["status"] == "Fail")
                status_parts.append(f"{'❌' if fails else '✅'} Validation ({len(val_results)} rules)")
            else:
                status_parts.append("⬜ Validation (not run)")
            if fraud:
                status_parts.append(f"🔍 Fraud score: **{fraud.get('fraud_score', 'N/A')}** — {fraud.get('risk_level', 'N/A')}")
            else:
                status_parts.append("⬜ Fraud scoring (not run)")
            st.markdown("  &nbsp;·&nbsp;  ".join(status_parts), unsafe_allow_html=True)

        with dl_col:
            # Generate report HTML and offer download
            report_html = _report_html(
                filename, meta, doc_type, extracted, val_results, fraud
            )
            safe_name = claim_ref.replace("/", "-") if claim_ref != "N/A" else filename
            st.download_button(
                label="⬇ Download Report",
                data=report_html.encode("utf-8"),
                file_name=f"claim_report_{safe_name}.html",
                mime="text/html",
                key=f"dl_{filename}",
                use_container_width=True,
            )
            try:
                workbook = _report_excel_bytes(
                    filename, meta, doc_type, extracted, val_results, fraud
                )
                st.download_button(
                    label="Download Excel",
                    data=workbook,
                    file_name=f"claim_report_{safe_name}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key=f"dl_excel_{filename}",
                    use_container_width=True,
                )
            except RuntimeError as exc:
                st.caption(str(exc))

        # ── Inline preview ────────────────────────────────────────────────────
        with st.expander("👁 Preview report inline"):
            st.components.v1.html(report_html, height=800, scrolling=True)

        # ── Quick summary cards ───────────────────────────────────────────────
        st.markdown("<br>", unsafe_allow_html=True)
        sc1, sc2, sc3 = st.columns(3)
        with sc1:
            section_header("Extracted Fields")
            filled = sum(1 for v in extracted.values() if v not in (None, "", []))
            total  = len(extracted)
            st.markdown(
                metric_card("Fields Extracted", f"{filled} / {total}",
                            f"{filled/total*100:.0f}% complete" if total else "—", filled == total),
                unsafe_allow_html=True,
            )
        with sc2:
            section_header("Validation")
            if val_results:
                passes   = sum(1 for r in val_results if r["status"] == "Pass")
                fails    = sum(1 for r in val_results if r["status"] == "Fail")
                warnings = sum(1 for r in val_results if r["status"] == "Warning")
                verdict  = "Pass" if fails == 0 else "Fail"
                st.markdown(
                    metric_card("Rules Run", str(len(val_results)),
                                f"{passes}✓ {fails}✗ {warnings}⚠", fails == 0),
                    unsafe_allow_html=True,
                )
                st.markdown(f"Overall: {badge_html(verdict)}", unsafe_allow_html=True)
            else:
                st.caption("Validation not yet run.")
        with sc3:
            section_header("Fraud Score")
            if fraud and "fraud_score" in fraud:
                score = fraud["fraud_score"]
                risk  = fraud.get("risk_level", "Unknown")
                low   = score < 30
                st.markdown(
                    metric_card("Fraud Score", f"{score} / 100",
                                risk, low),
                    unsafe_allow_html=True,
                )
                st.markdown(fraud_bar_html(score), unsafe_allow_html=True)
            else:
                st.caption("Fraud scoring not yet run.")

        st.markdown("</div>", unsafe_allow_html=True)
        st.markdown("<br>", unsafe_allow_html=True)

    # ── Bulk download: all reports in one zip ─────────────────────────────────
    if len(ready) > 1:
        st.markdown("---")
        section_header("Bulk Export")
        import zipfile, io
        zip_buf = io.BytesIO()
        with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
            for filename in ready:
                meta      = st.session_state.get(f"meta_{filename}", {})
                doc_type  = st.session_state.get(f"cls_{filename}", "Unknown")
                extracted = st.session_state.get(f"ext_{filename}", {})
                val_results = st.session_state.get(f"val_{filename}", [])
                fraud     = st.session_state.get(f"fraud_{filename}", {})
                html_content = _report_html(filename, meta, doc_type, extracted, val_results, fraud)
                safe = meta.get("claim_ref", filename).replace("/", "-")
                zf.writestr(f"claim_report_{safe}.html", html_content)
        zip_buf.seek(0)
        st.download_button(
            label=f"⬇ Download All {len(ready)} Reports (.zip)",
            data=zip_buf.read(),
            file_name="marvelai_claim_reports.zip",
            mime="application/zip",
            type="primary",
        )

# ──────────────────────────────────────────────────────────────────────────────
# SIDEBAR NAVIGATION
# ──────────────────────────────────────────────────────────────────────────────
def render_sidebar() -> str:
    with st.sidebar:
        # Logo / Brand
        st.markdown("""
        <div style="padding:1.45rem 1rem 1.05rem;">
            <div style="display:flex;align-items:center;gap:0.65rem;">
                <div style="width:34px;height:34px;border-radius:8px;background:#C98A2E;
                            color:#172033;font-weight:850;display:flex;align-items:center;
                            justify-content:center;letter-spacing:0;">M</div>
                <div style="font-size:1.2rem;font-weight:800;letter-spacing:0;color:#FFFFFF;">
                    MarvelAI
                </div>
            </div>
            <div style="font-size:0.68rem;color:rgba(255,255,255,0.5);letter-spacing:0.13em;
                        text-transform:uppercase;margin-top:0.65rem;">
                Travel PA Claims Portal
            </div>
        </div>
        <hr style="border:none;border-top:1px solid rgba(255,255,255,0.1);margin:0 1rem 0.75rem;">
        """, unsafe_allow_html=True)

        # LLM status indicator
        llm_status = get_ollama_status()
        llm_ok = bool(llm_status["online"] and llm_status["model"])

        if llm_ok:
            st.markdown(
                '<div style="padding:0 1rem 0.5rem;">'
                '<span style="font-size:0.75rem;color:rgba(255,255,255,0.55);">'
                f'● <span style="color:#4ADE9F;">Online</span> &nbsp;·&nbsp; {active_model_label()}</span></div>',
                unsafe_allow_html=True,
            )
        else:
            st.markdown(
                '<div style="padding:0 1rem 0.5rem;">'
                '<span style="font-size:0.75rem;color:rgba(255,255,255,0.55);">'
                '● <span style="color:#F87171;">Offline</span> &nbsp;·&nbsp; Ollama/model unavailable</span></div>',
                unsafe_allow_html=True,
            )

        st.markdown("<br>", unsafe_allow_html=True)

        # Navigation
        pages = {
            "Dashboard"                  : "dashboard",
            "Claim Ingestion"            : "ingestion",
            "Processing & Extraction"    : "processing",
            "Validation"                 : "validation",
            "AI Fraud Scoring"           : "fraud",
            "Claim Report"               : "report",
        }

        selected_label = st.radio(
            "Navigation",
            list(pages.keys()),
            label_visibility="collapsed",
        )

        st.markdown(
            "<hr style='border:none;border-top:1px solid rgba(255,255,255,0.1);"
            "margin:0.75rem 0;'>",
            unsafe_allow_html=True,
        )

        # Session info
        doc_count = len([k for k in st.session_state if k.startswith("ocr_")])
        st.markdown(
            f'<div style="padding:0 0.25rem 0.5rem;font-size:0.75rem;'
            f'color:rgba(255,255,255,0.45);">'
            f'📂 &nbsp;{doc_count} document(s) in session</div>',
            unsafe_allow_html=True,
        )

        if st.button("Clear Session", use_container_width=True):
            for k in list(st.session_state.keys()):
                del st.session_state[k]
            st.rerun()

        # Version footer
        st.markdown(
            '<div style="font-size:0.65rem;color:rgba(255,255,255,0.28);'
            'margin-top:2rem;text-align:center;padding-bottom:1rem;">'
            f'v1.1 &nbsp;·&nbsp; {active_model_label()} via Ollama</div>',
            unsafe_allow_html=True,
        )

        return pages[selected_label]


# ──────────────────────────────────────────────────────────────────────────────
# MAIN ENTRYPOINT
# ──────────────────────────────────────────────────────────────────────────────
def main() -> None:
    inject_css()

    # Generate (or retrieve cached) mock claims data
    if "mock_claims_df" not in st.session_state:
        st.session_state["mock_claims_df"] = generate_mock_claims(120)

    # Render sidebar and get active page
    active_page = render_sidebar()

    # Route to the selected page
    if active_page == "dashboard":
        page_dashboard(st.session_state["mock_claims_df"])
    elif active_page == "ingestion":
        page_ingestion()
    elif active_page == "processing":
        page_processing()
    elif active_page == "validation":
        page_validation()
    elif active_page == "fraud":
        page_fraud_scoring()
    elif active_page == "report":
        page_report()


if __name__ == "__main__":
    main()
